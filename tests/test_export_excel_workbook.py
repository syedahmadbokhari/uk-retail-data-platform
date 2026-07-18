"""
Unit tests for src/analysis/export_excel_workbook.py  build_workbook()
Uses mocks and synthetic data — no database connection required.
"""
import os
import numpy as np
import pandas as pd
import pytest
import openpyxl
from unittest.mock import patch, MagicMock

from src.analysis.export_excel_workbook import build_workbook

# ── Synthetic fixtures ────────────────────────────────────────────────────────

_RAW_DF = pd.DataFrame({
    "product_id":    ["P001", "P002", "P003", "P004"],
    "product_name":  ["Shoe A", "Shoe B", "Shoe C", "Shoe D"],
    "brand":         ["Adidas", "Nike", "Adidas", "Nike"],
    "listing_price": [100.0, 150.0, 80.0, 120.0],
    "discount":      [0.0, 0.3, 0.0, 0.4],
    "revenue":       [500.0, 800.0, 300.0, 900.0],
})

_BRAND_DF = pd.DataFrame({
    "brand":             ["Adidas", "Nike"],
    "total_revenue":     [800.0, 1700.0],
    "product_count":     [2, 2],
    "revenue_share_pct": [32.0, 68.0],
})

# Manufactured groups with a known, real (non-degenerate) difference so the
# hypothesis test underneath runs on genuine data rather than a fluke.
_DISCOUNTED = pd.Series(np.random.default_rng(42).normal(loc=900, scale=50, size=30))
_FULL_PRICE = pd.Series(np.random.default_rng(7).normal(loc=400, scale=50, size=30))


def _build(tmp_path, output_name="test_workbook.xlsx"):
    output_path = str(tmp_path / output_name)
    mock_conn = MagicMock()
    mock_conn.__enter__ = MagicMock(return_value=mock_conn)
    mock_conn.__exit__ = MagicMock(return_value=False)

    with patch("src.analysis.export_excel_workbook.get_connection", return_value=mock_conn), \
         patch("src.analysis.export_excel_workbook.pd.read_sql", side_effect=[_RAW_DF, _BRAND_DF]), \
         patch("src.analysis.export_excel_workbook.get_revenue_groups", return_value=(_DISCOUNTED, _FULL_PRICE)):
        result_path = build_workbook(output_path=output_path)

    return result_path


# ── Runs without error, produces a valid file ─────────────────────────────────

def test_build_workbook_runs_without_error_and_creates_file(tmp_path):
    result_path = _build(tmp_path)
    assert os.path.exists(result_path)
    assert os.path.getsize(result_path) > 0


def test_build_workbook_returns_a_valid_xlsx(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)   # raises if the zip/XML is invalid
    assert wb is not None


# ── Expected sheet names and order ────────────────────────────────────────────

def test_build_workbook_has_expected_sheet_names(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)
    assert wb.sheetnames == [
        "Summary", "Raw Data", "Revenue by Brand", "Discount Impact", "Pricing Opportunity Model",
    ]


def test_summary_sheet_is_first_tab(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)
    assert wb.sheetnames[0] == "Summary"


# ── Raw Data sheet — real row-level data, not aggregated ──────────────────────

def test_raw_data_sheet_contains_all_synthetic_rows(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)
    ws = wb["Raw Data"]
    assert ws.max_row == len(_RAW_DF) + 1   # header + rows
    assert [c.value for c in ws[1]] == list(_RAW_DF.columns)


# ── Native charts present on the aggregate sheets ─────────────────────────────

def test_revenue_by_brand_sheet_has_a_native_chart(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)
    assert len(wb["Revenue by Brand"]._charts) == 1


def test_discount_impact_sheet_has_a_native_chart(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)
    assert len(wb["Discount Impact"]._charts) == 1


# ── Discount Impact sheet carries the real statistical result ────────────────

def test_discount_impact_sheet_has_statistical_comment(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path)
    ws = wb["Discount Impact"]
    comment_cells = [c for row in ws.iter_rows() for c in row if c.comment is not None]
    assert len(comment_cells) == 1
    assert "p =" in comment_cells[0].comment.text or "p-value" in comment_cells[0].comment.text.lower() \
        or "statistically" in comment_cells[0].comment.text.lower()


# ── Pricing Opportunity Model sheet — genuine Excel formulas ─────────────────

def test_pricing_model_uses_real_formulas_not_static_values(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path, data_only=False)
    ws = wb["Pricing Opportunity Model"]

    formula_cells = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    assert len(formula_cells) >= 3


def test_pricing_model_inputs_reflect_real_group_stats(tmp_path):
    result_path = _build(tmp_path)
    wb = openpyxl.load_workbook(result_path, data_only=False)
    ws = wb["Pricing Opportunity Model"]

    values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)}
    assert values["Full-price product count"] == len(_FULL_PRICE)
    assert values["Full-price median revenue per product (£)"] == pytest.approx(float(_FULL_PRICE.median()))
    assert values["Discounted median revenue per product (£)"] == pytest.approx(float(_DISCOUNTED.median()))
